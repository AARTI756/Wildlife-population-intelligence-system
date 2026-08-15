import os
import csv
import time
from datetime import datetime, timedelta, timezone
from typing import Dict, Any, List, Optional
from sqlalchemy.orm import Session
from sqlalchemy import func


def format_ist_timestamp(dt: datetime = None) -> str:
    """Convert a UTC datetime (naive or aware) to IST (Asia/Kolkata, UTC+05:30).
    Returns a clean string like '2026-08-14 19:55:04 IST'.
    If dt is None, uses the current UTC time.
    """
    IST = timezone(timedelta(hours=5, minutes=30))
    if dt is None:
        dt = datetime.now(timezone.utc)
    # Ensure the datetime is timezone-aware
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    ist_dt = dt.astimezone(IST)
    return ist_dt.strftime("%Y-%m-%d %H:%M:%S IST")

from app.models.report import ReportHistory, ReportAuditLog
from app.models.monitoring import Survey, MonitoringSite, CameraTrap, AudioSensor
from app.models.observation import Observation

# Import from other services to avoid calculations duplication
from app.services.population_estimation import get_population_overview, get_species_metrics
from app.services.biodiversity_analytics import get_biodiversity_overview, get_species_composition, get_endangered_summary, get_relative_abundance, get_species_profile_map
from app.services.habitat_intelligence import get_habitat_overview, get_habitat_classification, get_vegetation_analysis
from app.services.conservation_recommendations import get_conservation_overview, get_conservation_priorities, get_actionable_recommendations
from app.services.health_scoring import get_health_overview, get_health_breakdown, get_health_alerts
from app.services.executive_dashboard import get_executive_overview, get_executive_population_trends

# PDF & Excel formatting imports
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Try to register Arial font to handle rupee sign (₹) on Windows
pdf_font = "Helvetica"
pdf_font_bold = "Helvetica-Bold"

for font_path in [
    "C:\\Windows\\Fonts\\arial.ttf",
    "C:\\Windows\\Fonts\\calibri.ttf",
    "C:\\Windows\\Fonts\\segoeui.ttf",
    "arial.ttf"
]:
    if os.path.exists(font_path):
        try:
            pdfmetrics.registerFont(TTFont('Arial', font_path))
            pdf_font = "Arial"
            break
        except Exception:
            pass

for font_path in [
    "C:\\Windows\\Fonts\\arialbd.ttf",
    "C:\\Windows\\Fonts\\calibrib.ttf",
    "C:\\Windows\\Fonts\\segoeuib.ttf",
    "arialbd.ttf"
]:
    if os.path.exists(font_path):
        try:
            pdfmetrics.registerFont(TTFont('Arial-Bold', font_path))
            pdf_font_bold = "Arial-Bold"
            break
        except Exception:
            pass

class NumberedCanvas(canvas.Canvas):
    """Custom canvas that computes total pages dynamically and adds header/footer lines."""
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_decorations(num_pages)
            super().showPage()
        super().save()

    def draw_page_decorations(self, page_count):
        self.saveState()
        self.setFont(pdf_font, 8)
        self.setFillColor(colors.HexColor("#475569"))
        
        # Header (Only on page 2 and later)
        if self._pageNumber > 1:
            self.drawString(54, 755, "Wildlife Population Intelligence System (WPIS) - Detailed Report")
            self.setStrokeColor(colors.HexColor("#E2E8F0"))
            self.setLineWidth(0.5)
            self.line(54, 748, 558, 748)
            
        # Footer (On all pages)
        self.setStrokeColor(colors.HexColor("#E2E8F0"))
        self.setLineWidth(0.5)
        self.line(54, 48, 558, 48)
        
        page_text = f"Page {self._pageNumber} of {page_count}"
        self.drawRightString(558, 36, page_text)
        self.drawString(54, 36, "CONFIDENTIAL - WPIS Official Forest Department Document")
        self.restoreState()


def parse_date(date_str: Optional[str]) -> Optional[datetime]:
    """Defensive date parser supporting ISO and YYYY-MM-DD strings."""
    if not date_str:
        return None
    try:
        cleaned = date_str.replace("Z", "+00:00")
        return datetime.fromisoformat(cleaned)
    except Exception:
        try:
            return datetime.strptime(date_str, "%Y-%m-%d")
        except Exception:
            return None

def compile_report_dataset(db: Session, report_type: str, filters: Dict[str, Any]) -> Dict[str, Any]:
    cleaned_filters = {k: v for k, v in filters.items() if v is not None}
    
    # Remap UI filter keys to service parameters
    if "start_date" in cleaned_filters:
        cleaned_filters["date_from"] = cleaned_filters.pop("start_date")
    if "end_date" in cleaned_filters:
        cleaned_filters["date_to"] = cleaned_filters.pop("end_date")
        
    # Standardize dates and IDs to ensure database comparisons match expected types
    if "date_from" in cleaned_filters:
        cleaned_filters["date_from"] = parse_date(cleaned_filters["date_from"])
    if "date_to" in cleaned_filters:
        cleaned_filters["date_to"] = parse_date(cleaned_filters["date_to"])
        
    for id_key in ["survey_id", "site_id"]:
        if id_key in cleaned_filters and cleaned_filters[id_key] is not None:
            try:
                cleaned_filters[id_key] = int(cleaned_filters[id_key])
            except (ValueError, TypeError):
                pass

    # Build a sanitised filter dict that only passes keys accepted by service
    # functions (i.e. kwargs of get_filtered_observations / _apply_filters).
    # This prevents TypeError when UI-only keys like conservation_status are
    # present in the request payload.
    ALLOWED_SERVICE_KEYS = {"survey_id", "site_id", "species", "habitat",
                            "date_from", "date_to", "protected_area", "state"}
    service_filters = {k: v for k, v in cleaned_filters.items() if k in ALLOWED_SERVICE_KEYS}

    dataset = {
        "generated_at": format_ist_timestamp(),
        "report_type": report_type,
        "filters": cleaned_filters
    }
    
    if report_type == "Wildlife Survey Report":
        # Survey base numbers
        survey_query = db.query(Survey)
        site_query = db.query(MonitoringSite)
        obs_query = db.query(Observation)
        
        if "survey_id" in cleaned_filters:
            survey_query = survey_query.filter(Survey.id == cleaned_filters["survey_id"])
            obs_query = obs_query.filter(Observation.survey_id == cleaned_filters["survey_id"])
        if "site_id" in cleaned_filters:
            survey_query = survey_query.filter(Survey.monitoring_site_id == cleaned_filters["site_id"])
            obs_query = obs_query.filter(Observation.monitoring_site_id == cleaned_filters["site_id"])
            
        surveys_list = survey_query.order_by(Survey.date.desc()).all()
        sites_count = site_query.count()
        observations_count = obs_query.count()
        cameras_count = db.query(CameraTrap).count()
        audios_count = db.query(AudioSensor).count()
        
        dataset["summary"] = {
            "Total Surveys": len(surveys_list),
            "Monitoring Sites": sites_count,
            "Total Observations": observations_count,
            "Active Cameras": cameras_count,
            "Active Audio Sensors": audios_count
        }
        dataset["data_table"] = [
            {"Survey ID": s.id, "Survey Name": s.name, "Date": str(s.date), "Device": s.monitoring_device, "Area": "Protected" if s.protected_area else "Buffer"}
            for s in surveys_list
        ]
        
    elif report_type == "Species Population Report":
        # Separate conservation_status since services don't consume it directly
        target_status = cleaned_filters.get("conservation_status")
        service_filters = {k: v for k, v in cleaned_filters.items() if k != "conservation_status"}
        
        overview = get_population_overview(db, **service_filters)
        species_metrics = get_species_metrics(db, **service_filters)
        
        if target_status:
            # Filter metrics based on IUCN status
            species_metrics = [m for m in species_metrics if (m.get("iucn_status") or "").strip().lower() == target_status.strip().lower()]
            total_est = sum(m.get("estimated_population", 0) for m in species_metrics)
            richness = len(species_metrics)
            dataset["summary"] = {
                "Total Population Estimate": total_est,
                "Unique Species Count": richness,
                "Avg Density (ind/km²)": f"{overview.get('average_density', 0.0):.2f}",
                "Coverage Factor (%)": f"{overview.get('average_observation_coverage', 0.0):.1f}%"
            }
        else:
            dataset["summary"] = {
                "Total Population Estimate": overview.get("total_estimated_population", 0),
                "Unique Species Count": overview.get("total_species_richness", 0),
                "Avg Density (ind/km²)": f"{overview.get('average_density', 0.0):.2f}",
                "Coverage Factor (%)": f"{overview.get('average_observation_coverage', 0.0):.1f}%"
            }
            
        dataset["data_table"] = [
            {
                "Species Common Name": m.get("species_name", "—"),
                "Scientific Name": m.get("scientific_name", "—"),
                "Taxon Class": m.get("taxon_class", "—"),
                "Observations": m.get("observation_count", 0),
                "Est. Population": m.get("estimated_population", 0),
                "Detection Freq (%)": f"{m.get('detection_frequency', 0.0):.1f}%",
                "IUCN Status": m.get("iucn_status", "—")
            }
            for m in species_metrics
        ]
        
    elif report_type == "Biodiversity Report":
        target_status = cleaned_filters.get("conservation_status")
        service_filters = {k: v for k, v in cleaned_filters.items() if k != "conservation_status"}
        
        overview = get_biodiversity_overview(db, **service_filters)
        comp = get_relative_abundance(db, **service_filters)
        endangered = get_endangered_summary(db, **service_filters)
        
        species_map = get_species_profile_map(db)
        
        if target_status:
            # Safe IUCN-status filter: guard against None species_name or missing profile
            filtered_comp = []
            for c in comp:
                sp_name = c.get("species_name") or ""
                profile = species_map.get(sp_name.lower().strip()) if sp_name else None
                iucn = (profile[3] if (profile and len(profile) > 3) else "") or ""
                if iucn.strip().lower() == target_status.strip().lower():
                    filtered_comp.append(c)
            comp = filtered_comp
            endangered = [
                e for e in endangered
                if (e.get("iucn_status") or "").strip().lower() == target_status.strip().lower()
            ]

        dataset["summary"] = {
            "Shannon Index (H')": f"{overview.get('shannon_diversity_index', 0.0):.3f}",
            "Simpson Index (D)": f"{overview.get('simpson_diversity_index', 0.0):.3f}",
            "Evenness (E)": f"{overview.get('species_evenness', 0.0):.3f}",
            "Species Richness": overview.get("species_richness", 0)
        }
        dataset["data_table"] = [
            {
                "Species Name": c.get("species_name", "—"),
                "Scientific Name": c.get("scientific_name", "—"),
                "Observations": c.get("observation_count", 0),
                "Relative Abundance (%)": f"{c.get('relative_abundance_pct', 0.0):.2f}%"
            }
            for c in comp
        ]
        dataset["extra_table"] = [
            {
                "Threatened Species": e.get("species_name"),
                "IUCN Group": e.get("iucn_status"),
                # Use correct key: get_endangered_summary returns 'observation_count'
                "Observations Logged": e.get("observation_count")
            }
            for e in endangered
        ]
        
    elif report_type == "Habitat Assessment Report":
        overview = get_habitat_overview(db, **service_filters)
        classification = get_habitat_classification(db, **service_filters)
        veg = get_vegetation_analysis(db, **service_filters)
        
        if veg:
            avg_ndvi = sum(x["ndvi"] for x in veg) / len(veg)
            avg_ndvi_str = f"{avg_ndvi:.2f}"
        else:
            avg_ndvi_str = "No Data"
            
        dataset["summary"] = {
            "Ecosystem Quality Index": f"{overview.get('habitat_quality_score', 0.0):.1f}/100",
            "Human Disturbance Factor": f"{overview.get('human_disturbance', 0.0):.1f}/100",
            "Average NDVI Indicator": avg_ndvi_str
        }
        dataset["data_table"] = [
            {
                "Biome Classification": c.get("name") or "—",
                "Relative Coverage (%)": f"{c.get('value', 0.0):.1f}%",
                "Observations": c.get("observations", 0),
                "Anomaly Alert": "Low Coverage" if c.get("value", 0.0) < 5.0 else "Stable"
            }
            for c in classification
        ]
        
    elif report_type == "Conservation Report":
        overview = get_conservation_overview(db, **service_filters)
        priorities = get_conservation_priorities(db, **service_filters)
        actions = get_actionable_recommendations(db, **service_filters)
        
        import re
        total_action_cost = 0.0
        for a in actions:
            cost_str = a.get("estimated_cost", "₹0")
            cleaned_cost = re.sub(r'[^\d.]', '', cost_str)
            try:
                numeric_cost = float(cleaned_cost) if cleaned_cost else 0.0
                total_action_cost += numeric_cost
            except ValueError:
                pass
                
        health_overview = get_health_overview(db, **service_filters)
        ecosystem_health = health_overview.get("overallScore", 0.0)
        
        dataset["summary"] = {
            "Active Recommendations": len(actions),
            "Ecosystem Health Score": f"{ecosystem_health:.1f}/100",
            "Total Required Funding": f"₹{total_action_cost:,.2f}" if total_action_cost > 0 else "₹0.00"
        }
        
        dataset["data_table"] = []
        for a in actions:
            cost_str = a.get("estimated_cost", "₹0")
            cleaned_cost = re.sub(r'[^\d.]', '', cost_str)
            try:
                cost_val = float(cleaned_cost) if cleaned_cost else 0.0
                formatted_cost = f"₹{cost_val:,.2f}"
            except ValueError:
                formatted_cost = "₹0.00"
                
            dataset["data_table"].append({
                "Action Description": a.get("title"),
                "Priority Status": a.get("priority"),
                "Department Owner": a.get("department", "Forest Dept"),
                "Completion Time": a.get("completion_time", "30 Days"),
                "Estimated Cost": formatted_cost,
                "Expected Impact": a.get("expected_impact", "N/A")
            })
        
    elif report_type == "Wildlife Health Report":
        overview = get_health_overview(db, **service_filters)
        breakdown = get_health_breakdown(db, **service_filters)
        
        dataset["summary"] = {
            "Overall Health Score": f"{overview.get('overallScore', 0.0):.1f}/100",
            "Ecosystem Rating": overview.get("statusName", "Moderate Concern"),
            "Calculations Framework": "Weighted Index Model (30% Species Diversity, 25% Population Stability, 20% Habitat Quality, 15% Endangered Species Status, 10% Environmental Conditions)"
        }
        
        def _get_pillar_status(score):
            if score >= 75: return "Healthy"
            if score >= 60: return "Moderate Concern"
            return "Critical"
            
        dataset["data_table"] = [
            {
                "Ecosystem Index Pillar": b.get("name"),
                "Assigned Weighted Score": f"{b.get('value', 0.0):.1f}/100",
                "Assigned Weight (%)": f"{b.get('weight', 0.0):.0f}%",
                "Ecosystem Rating": _get_pillar_status(b.get("value", 0.0))
            }
            for b in breakdown
        ]
        
    elif report_type == "Executive Summary Report":
        overview = get_executive_overview(db, **service_filters)
        health_data = get_health_overview(db, **service_filters)
        bio_data = get_biodiversity_overview(db, **service_filters)
        hab_data = get_habitat_overview(db, **service_filters)
        
        dataset["summary"] = {
            "Ecosystem Health Rating": f"{health_data.get('overallScore', 0.0):.1f}/100",
            "Shannon Index": f"{bio_data.get('shannon_diversity_index', 0.0):.3f}",
            "Habitat Quality Score": f"{hab_data.get('habitat_quality_score', 0.0):.1f}/100",
            "Total Live Observations": overview.get("metrics", {}).get("totalObservations", {}).get("value", "0")
        }
        dataset["data_table"] = [
            {
                "Ecosystem Index Pillar": "Overall Ecosystem Health",
                "Score Value": f"{health_data.get('overallScore', 0.0):.1f}/100",
                "Threshold Label": health_data.get("statusName", "Unknown")
            },
            {
                "Ecosystem Index Pillar": "Shannon Diversity Index",
                "Score Value": f"{bio_data.get('shannon_diversity_index', 0.0):.3f}",
                "Threshold Label": "Stable"
            },
            {
                "Ecosystem Index Pillar": "Habitat Suitability Index",
                "Score Value": f"{hab_data.get('habitat_quality_score', 0.0):.1f}/100",
                "Threshold Label": "Moderate"
            }
        ]
        
    else:
        dataset["summary"] = {"Event": "Empty dataset"}
        dataset["data_table"] = []
        
    return dataset


def generate_pdf_report(dataset: Dict[str, Any], filepath: str):
    """Write dynamic reportlab PDF using canvas template decorators."""
    doc = SimpleDocTemplate(
        filepath,
        pagesize=letter,
        leftMargin=54,
        rightMargin=54,
        topMargin=72,
        bottomMargin=72
    )
    
    styles = getSampleStyleSheet()
    
    # Custom colors
    primary_color = colors.HexColor("#0F766E") # Teal-700
    slate_color = colors.HexColor("#1E293B")
    
    # Modify styles
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName=pdf_font_bold,
        fontSize=22,
        textColor=primary_color,
        spaceAfter=15
    )
    
    h2_style = ParagraphStyle(
        'DocH2',
        parent=styles['Heading2'],
        fontName=pdf_font_bold,
        fontSize=13,
        textColor=slate_color,
        spaceBefore=15,
        spaceAfter=8
    )
    
    body_style = ParagraphStyle(
        'DocBody',
        parent=styles['Normal'],
        fontName=pdf_font,
        fontSize=10,
        leading=14,
        textColor=colors.HexColor("#334155")
    )
    
    bold_body_style = ParagraphStyle(
        'DocBoldBody',
        parent=body_style,
        fontName=pdf_font_bold
    )
    
    story = []
    
    # Title
    story.append(Paragraph(dataset["report_type"], title_style))
    
    # Metadata Block
    meta_data = [
        [Paragraph("Organization:", bold_body_style), Paragraph("Wildlife Population Intelligence System (WPIS)", body_style)],
        [Paragraph("Generated At:", bold_body_style), Paragraph(dataset["generated_at"], body_style)],
        [Paragraph("Filters Applied:", bold_body_style), Paragraph(str(dataset["filters"]) if dataset["filters"] else "None (All Sites)", body_style)]
    ]
    meta_table = Table(meta_data, colWidths=[100, 400])
    meta_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('TOPPADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 20))
    
    # Executive Summary Cards Table
    story.append(Paragraph("Executive Summary Metrics", h2_style))
    summary_data = []
    for k, v in dataset.get("summary", {}).items():
        summary_data.append([Paragraph(str(k), bold_body_style), Paragraph(str(v), body_style)])
        
    summary_table = Table(summary_data, colWidths=[200, 300])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#F8FAFC")),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E2E8F0")),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('PADDING', (0,0), (-1,-1), 8),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 20))
    
    # Detailed Dataset Table
    story.append(Paragraph("Detailed Analytical Breakdown", h2_style))
    table_rows = dataset.get("data_table", [])
    if not table_rows:
        story.append(Paragraph("No tabular dataset logs recorded for this layout scope.", body_style))
    else:
        # Extract headers from first row keys
        headers = list(table_rows[0].keys())
        
        pdf_table_data = []
        # Header Row
        pdf_table_data.append([Paragraph(h, ParagraphStyle('HeaderStyle', parent=bold_body_style, textColor=colors.white)) for h in headers])
        
        # Data Rows
        for row in table_rows:
            pdf_table_data.append([Paragraph(str(row.get(h, "")), body_style) for h in headers])
            
        # Calculate custom column widths based on header values to prevent text wrapping or squeezing overlaps (total printable width = 504)
        num_cols = len(headers)
        col_widths = []
        if "Species Common Name" in headers:
            col_widths = [100, 80, 80, 54, 60, 80, 50]  # Total = 504 (7 columns)
        elif "Species Name" in headers:
            col_widths = [150, 150, 100, 104]    # Total = 504 (4 columns)
        elif "Biome Classification" in headers:
            col_widths = [150, 120, 114, 120]    # Total = 504 (4 columns)
        elif "Action Description" in headers:
            col_widths = [154, 70, 70, 70, 70, 70] # Total = 504 (6 columns)
        elif "Ecosystem Index Pillar" in headers:
            col_widths = [184, 160, 160] if num_cols == 3 else [200, 104, 100, 100]
        else:
            col_widths = [504 / num_cols] * num_cols

        # Adjust dynamically if headers length doesn't match default mapping sizes
        if len(col_widths) != num_cols:
            col_widths = [504 / num_cols] * num_cols
            
        detail_table = Table(pdf_table_data, colWidths=col_widths)
        detail_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), primary_color),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#CBD5E1")),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor("#F8FAFC")]),
        ]))
        story.append(detail_table)
        
    # Extra table (e.g. threatened list in Biodiversity)
    extra_rows = dataset.get("extra_table", [])
    if extra_rows:
        story.append(Spacer(1, 20))
        story.append(Paragraph("Threatened Species Log Details", h2_style))
        extra_headers = list(extra_rows[0].keys())
        
        extra_pdf_data = []
        extra_pdf_data.append([Paragraph(h, ParagraphStyle('HeaderStyle2', parent=bold_body_style, textColor=colors.white)) for h in extra_headers])
        for row in extra_rows:
            extra_pdf_data.append([Paragraph(str(row.get(h, "")), body_style) for h in extra_headers])
            
        num_cols = len(extra_headers)
        col_width = 504 / num_cols
        extra_table = Table(extra_pdf_data, colWidths=[col_width]*num_cols)
        extra_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#B91C1C")), # Red-700 header
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#FCA5A5")),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor("#FEF2F2")]),
        ]))
        story.append(extra_table)
        
    # Build Document
    doc.build(story, canvasmaker=NumberedCanvas)


def generate_excel_report(dataset: Dict[str, Any], filepath: str):
    """Write dynamic openpyxl workbook with zebra styling and auto column fits."""
    wb = openpyxl.Workbook()
    
    # 1. Overview sheet
    ws1 = wb.active
    ws1.title = "Overview"
    ws1.views.sheetView[0].showGridLines = True
    
    # Title formatting
    ws1["A1"] = dataset["report_type"]
    ws1["A1"].font = Font(name="Arial", size=14, bold=True, color="0F766E")
    
    # Metadata
    ws1["A3"] = "Organization"
    ws1["B3"] = "Wildlife Population Intelligence System (WPIS)"
    ws1["A4"] = "Generated At"
    ws1["B4"] = dataset["generated_at"]
    ws1["A5"] = "Applied Filters"
    ws1["B5"] = str(dataset["filters"]) if dataset["filters"] else "None (All Sites)"
    
    for row in range(3, 6):
        ws1[f"A{row}"].font = Font(name="Arial", size=10, bold=True)
        ws1[f"B{row}"].font = Font(name="Arial", size=10)
        
    # Summary KPI card block
    ws1["A8"] = "Executive Summary Indicators"
    ws1["A8"].font = Font(name="Arial", size=11, bold=True, color="1E293B")
    
    row_idx = 9
    for k, v in dataset.get("summary", {}).items():
        ws1.cell(row=row_idx, column=1, value=k).font = Font(name="Arial", size=10, bold=True)
        ws1.cell(row=row_idx, column=2, value=v).font = Font(name="Arial", size=10)
        row_idx += 1
        
    # Format Summary Table
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    for r in range(9, row_idx):
        ws1.cell(row=r, column=1).border = thin_border
        ws1.cell(row=r, column=2).border = thin_border
        ws1.cell(row=r, column=1).fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        
    # Auto-adjust column widths for Overview
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # 2. Detailed Dataset sheet
    table_rows = dataset.get("data_table", [])
    if table_rows:
        ws2 = wb.create_sheet(title="Detailed Analytical Breakdown")
        ws2.views.sheetView[0].showGridLines = True
        
        headers = list(table_rows[0].keys())
        ws2.append(headers)
        
        # Freeze headers row
        ws2.freeze_panes = "A2"
        
        # Style Header
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
        for col_idx in range(1, len(headers) + 1):
            cell = ws2.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        # Append data & set zebra borders
        zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        for r_idx, row in enumerate(table_rows, start=2):
            ws2.append(list(row.values()))
            for col_idx in range(1, len(headers) + 1):
                cell = ws2.cell(row=r_idx, column=col_idx)
                cell.border = thin_border
                cell.font = Font(name="Arial", size=10)
                if r_idx % 2 == 1:
                    cell.fill = zebra_fill
                    
        # Auto-adjust column widths for Details
        for col in ws2.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws2.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    # Extra table (e.g. threatened list in Biodiversity)
    extra_rows = dataset.get("extra_table", [])
    if extra_rows:
        ws3 = wb.create_sheet(title="Extra Details")
        ws3.views.sheetView[0].showGridLines = True
        extra_headers = list(extra_rows[0].keys())
        ws3.append(extra_headers)
        ws3.freeze_panes = "A2"
        
        # Red headers for threatened lists
        red_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        red_fill = PatternFill(start_color="B91C1C", end_color="B91C1C", fill_type="solid")
        for col_idx in range(1, len(extra_headers) + 1):
            cell = ws3.cell(row=1, column=col_idx)
            cell.font = red_font
            cell.fill = red_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        # Append extra data
        for r_idx, row in enumerate(extra_rows, start=2):
            ws3.append(list(row.values()))
            for col_idx in range(1, len(extra_headers) + 1):
                cell = ws3.cell(row=r_idx, column=col_idx)
                cell.border = thin_border
                cell.font = Font(name="Arial", size=10)
                
        # Auto column dimensions
        for col in ws3.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws3.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    wb.save(filepath)


def generate_csv_report(dataset: Dict[str, Any], filepath: str):
    """Write dynamic flat CSV table using Python csv module."""
    table_rows = dataset.get("data_table", [])
    if not table_rows:
        # Fallback to Summary keys if data_table is empty
        table_rows = [{"Summary Key": k, "Value": v} for k, v in dataset.get("summary", {}).items()]
        
    headers = list(table_rows[0].keys())
    with open(filepath, mode='w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
        # Write headers
        writer.writerow(headers)
        # Write data rows
        for row in table_rows:
            writer.writerow(list(row.values()))


def run_report_generation(db: Session, report_id: int, user_id: int):
    """Process report generation task in the background, logging duration and errors."""
    t0 = time.time()
    report = db.query(ReportHistory).filter(ReportHistory.id == report_id).first()
    if not report:
        return
        
    try:
        # Compile live dataset
        dataset = compile_report_dataset(db, report.report_type, report.filters_json or {})
        
        # Absolute secure output directory path
        reports_dir = os.path.join("uploads", "reports")
        os.makedirs(reports_dir, exist_ok=True)
        
        filepath = os.path.join(reports_dir, report.download_filename)
        
        # Generate target format
        if report.format == "PDF":
            generate_pdf_report(dataset, filepath)
        elif report.format == "XLSX":
            generate_excel_report(dataset, filepath)
        elif report.format == "CSV":
            generate_csv_report(dataset, filepath)
            
        duration_ms = int((time.time() - t0) * 1000)
        
        # Update run status details
        report.status = "Completed"
        report.execution_time_ms = duration_ms
        db.commit()
        
        # Write report generation audit log
        audit = ReportAuditLog(
            report_id=report_id,
            user_id=user_id,
            action="generated",
            report_type=report.report_type,
            format=report.format
        )
        db.add(audit)
        db.commit()
        
    except Exception as e:
        print(f"Report Generation Failed: {e}")
        report.status = "Failed"
        db.commit()


def generate_wildlife_monitoring_report(
    filename: str,
    stored_filename: str,
    detections: List[Dict[str, Any]],
    biodiversity_metrics: Dict[str, Any],
    image_quality: Optional[Dict[str, Any]],
    processing_time_ms: float,
    survey_info: Dict[str, Any],
    prediction_type: str = "Image"
) -> Dict[str, Any]:
    """
    Compatibility helper generating structured monitoring report metadata for image/audio uploads.
    """
    total_detections = len(detections)
    high_confidence_detections = len([d for d in detections if d.get("confidence", 0) >= 0.70])
    
    report = {
        "report_type": "Telemetry Sighting Report",
        "format": "JSON",
        "generated_at": format_ist_timestamp(),
        "prediction_type": prediction_type,
        "input_file": filename,
        "stored_file": stored_filename,
        "processing_time_ms": round(processing_time_ms, 2),
        "survey_details": survey_info,
        "metrics": {
            "total_detections_count": total_detections,
            "high_confidence_count": high_confidence_detections,
            "biodiversity": biodiversity_metrics
        }
    }
    
    if image_quality:
        report["image_quality_metrics"] = image_quality
        
    return report
